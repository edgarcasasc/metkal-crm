import openpyxl

file_path = r"c:\Users\edyKzaz\metkal-crm\archivos-excel2\E-MK-26-0153 2.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb['Cálculos']
    
    print("Extracting formulas for the first measurement row (approx row 17):\n")
    for row in range(16, 20):
        print(f"--- Row {row} ---")
        for col in range(1, 25): # A to X
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None:
                print(f"{cell.coordinate}: {cell.value}")

    print("\nExtracting header details (Row 11-15):")
    for row in range(11, 15):
        for col in range(1, 20):
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None:
                print(f"{cell.coordinate}: {cell.value}")

except Exception as e:
    print("Error:", e)
