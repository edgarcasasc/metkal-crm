import msoffcrypto
import os
import io
import openpyxl

dir_path = r"c:\Users\edyKzaz\metkal-crm\archivos-excel"
files = ["M-MK-0497_1.xlsx", "DU-MK-0373.xlsx"]
passwords = ["m", "M"]

for f in files:
    print(f"\n{'='*50}\nFile: {f}")
    path = os.path.join(dir_path, f)
    with open(path, "rb") as file:
        office_file = msoffcrypto.OfficeFile(file)
        if office_file.is_encrypted():
            decrypted = None
            for pwd in passwords:
                try:
                    office_file.load_key(password=pwd)
                    decrypted_temp = io.BytesIO()
                    office_file.decrypt(decrypted_temp)
                    decrypted = decrypted_temp
                    break
                except Exception:
                    pass
            
            if decrypted:
                wb = openpyxl.load_workbook(decrypted, data_only=False)
                for sheet_name in ["Resultados I", "Anexo II", "Resultados"]:
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        print(f"\n--- Sheet: {sheet_name} ---")
                        for r in range(1, 100):
                            row_data = []
                            for c in range(1, 20):
                                cell = ws.cell(row=r, column=c)
                                if cell.value is not None:
                                    val_str = str(cell.value).replace('\n', ' ')
                                    row_data.append(f"{cell.coordinate}: {val_str}")
                            if row_data:
                                print(" | ".join(row_data))
            else:
                print("Decryption failed.")
