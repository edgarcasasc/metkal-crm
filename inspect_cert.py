import openpyxl
import sys
import os

def inspect_sheet(file_path, sheet_name):
    print(f"\n--- Inspecting {sheet_name} in {file_path} ---")
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found!")
        return

    ws = wb[sheet_name]
    
    # Let's collect all cells with values
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                # If it's a formula, let's also get the evaluated value if possible
                val_str = str(cell.value).replace('\n', ' ')
                print(f"{cell.coordinate}: {val_str[:150]}")

if __name__ == "__main__":
    files = ['archivos-excel/M-MK-0497_1.xlsx', 'archivos-excel/DU-MK-0373.xlsx']
    for f in files:
        if os.path.exists(f):
            inspect_sheet(f, "Certificado")
            inspect_sheet(f, "Resultados I")
        else:
            print(f"File not found: {f}")
